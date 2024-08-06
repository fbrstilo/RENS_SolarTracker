import sys
import os
import openpyxl
from   chirpstack_api import api
import grpc
import tracking as tr

class DeviceImportRecord:
    def __init__(self, dev_eui, device_profile_id, name, description):
        self.DevEUI          = dev_eui
        self.DeviceProfileID = device_profile_id
        self.Name            = name
        self.Description     = description

def get_device_import_list(file: str) -> list[DeviceImportRecord]:
    out = []
    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        print("open excel file error",e)
        return []

    sheet = wb.active
    rw_no = 0   
    rows = sheet.iter_rows(min_row=1,max_row=sheet.max_row)
    for a,b,c,d in rows:
        if rw_no > 0:
            print('importing rw:',rw_no,':',a.value,b.value,c.value,d.value)
            out.append(DeviceImportRecord(a.value,b.value,c.value,d.value))
        rw_no += 1
    
    return out

def import_devices(devices):
    server  = tr.chirpstack_server
    api_token = tr.api_token

    channel = grpc.insecure_channel(server)
    client  = api.DeviceServiceStub(channel)
    auth_token = [("authorization", "Bearer %s" % api_token)]
    nwk_key = '2B7E151628AED2A6ABF7158809CF4F3C'

    print('Input network key (leave empty for default)')
    user_input = input().strip()
    if(user_input != ''): nwk_key = user_input
    
    try:
        req_device = api.CreateDeviceRequest()
        for dev in devices:
            print('creating Device with DevEUI:',dev.DevEUI)
            req_device.device.dev_eui           = str(dev.DevEUI)
            req_device.device.name              = str(dev.Name)
            req_device.device.description       = str(dev.Description)
            req_device.device.application_id    = tr.app_id
            req_device.device.device_profile_id = str(dev.DeviceProfileID)
            req_device.device.is_disabled       = False                     
            resp = client.Create(req_device, metadata=auth_token)
        req_keys = api.CreateDeviceKeysRequest()
        for dev in devices:
            print('creating keys for device DevEUI:',dev.DevEUI)
            req_keys.device_keys.dev_eui = str(dev.DevEUI)
            req_keys.device_keys.nwk_key = nwk_key
            #req_keys.device_keys.app_key = str(dev.ApplicationKey)
            resp = client.CreateKeys(req_keys, metadata=auth_token)
    except  grpc.RpcError as e:
        if e.code() == grpc.StatusCode.INTERNAL:
            print('import error device',dev.DevEUI,' import aborted! Check Device my already exsist.')
        else:
            print('error:',type(e))

    return None

if __name__ == "__main__":
    if sys.argv.__contains__('--source'):
        index = sys.argv.index('--source') + 1
        if(sys.argv.__len__() < index + 1):
            print(f"Not enough arguments provided. Use: {sys.argv[0]} --path <path-to-your-table>")
            exit(0)
        path = sys.argv[index]
        if os.path.exists(path):
            dev_list = get_device_import_list(path)
            import_devices(dev_list)
        else:
            print(f'File {path} does not exist')
            exit(0)
    else:
        print(f'Use: python {sys.argv[0]} --source <path-to-your-table>')
        exit(0)